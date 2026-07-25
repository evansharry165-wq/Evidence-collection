"""Nightly runner — iterates every source, writes snapshots, updates the manifest workbook.

Invoked by .github/workflows/nightly-pull.yml (cron 03:00 UTC daily).

For each source:
    1. Import the source module from sources/
    2. Call pull() with a per-source timeout
    3. Write snapshot to data/YYYY-MM-DD/<id>.json and data/latest/<id>.json
    4. Record status in the manifest workbook
    5. Append a row to data/manifest_log.csv (append-only audit log)

Runs each source in isolation — one failure does not abort the run.
Exit code:  0 if all sources succeeded, 2 if one or more failed (workflow surfaces via email).
"""

import csv, importlib, json, pathlib, sys, traceback
from datetime import datetime, timezone

import openpyxl

sys.path.insert(0, str(pathlib.Path(__file__).parent))
sys.path.insert(0, str(pathlib.Path(__file__).parent / "sources"))
from pull_common import write_snapshot
from manifest_source_of_truth import SOURCES

ROOT = pathlib.Path(__file__).parent
MANIFEST = ROOT / "data" / "manifest.xlsx"
LOG = ROOT / "data" / "manifest_log.csv"


def utc_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def run_one(source: dict) -> dict:
    sid = source["id"]
    module_name = sid.replace("-", "_")
    result = {"id": sid, "status": "planned", "rows": 0, "bytes": 0, "live_url": "", "error": ""}
    try:
        mod = importlib.import_module(module_name)
    except ModuleNotFoundError:
        # Source declared in manifest but implementation not yet written — skip cleanly.
        result["status"] = "planned"
        return result
    try:
        payload = mod.pull()
        dated, size = write_snapshot(payload, ROOT)
        result["status"] = "live"
        result["rows"] = payload.get("rows", 0)
        result["bytes"] = size
        result["live_url"] = (
            f"https://raw.githubusercontent.com/evansharry165-wq/evidence-collection/main/"
            f"data/latest/{sid}.json"
        )
    except Exception as e:  # noqa: BLE001 — any failure is per-source, we log and continue
        result["status"] = "failed"
        result["error"] = f"{type(e).__name__}: {e}"
        print(f"FAILED {sid}: {result['error']}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
    return result


def update_manifest(results: list[dict]) -> None:
    if not MANIFEST.exists():
        print(f"WARN: {MANIFEST} not found — skipping workbook update", file=sys.stderr)
        return
    wb = openpyxl.load_workbook(MANIFEST)
    ws = wb["Sources"] if "Sources" in wb.sheetnames else wb.active
    # Build id -> row index map from column A
    id_to_row = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v:
            id_to_row[str(v).strip()] = r
    # Columns: 1 ID · 2 Cat · 3 Prov · 4 Endpoint · 5 Cadence · 6 Auth · 7 Cost · 8 Layer
    # 9 Status · 10 Last-pull-UTC · 11 Rows · 12 Bytes · 13 Live-URL · 14 Notes
    now = utc_iso()
    for r in results:
        row = id_to_row.get(r["id"])
        if not row:
            continue
        ws.cell(row=row, column=9, value=r["status"])
        ws.cell(row=row, column=10, value=now)
        ws.cell(row=row, column=11, value=r["rows"])
        ws.cell(row=row, column=12, value=r["bytes"])
        ws.cell(row=row, column=13, value=r["live_url"])
        if r["error"]:
            # Append error to notes without wiping the human-authored note
            existing = ws.cell(row=row, column=14).value or ""
            ws.cell(row=row, column=14, value=f"[last error {now}] {r['error']} · {existing}")
    wb.save(MANIFEST)


def append_log(results: list[dict]) -> None:
    exists = LOG.exists()
    with LOG.open("a", newline="") as f:
        w = csv.writer(f)
        if not exists:
            w.writerow(["run_utc", "source_id", "status", "rows", "bytes", "error"])
        now = utc_iso()
        for r in results:
            w.writerow([now, r["id"], r["status"], r["rows"], r["bytes"], r["error"]])


def main() -> int:
    results = [run_one(s) for s in SOURCES]
    update_manifest(results)
    append_log(results)
    failed = [r for r in results if r["status"] == "failed"]
    live = [r for r in results if r["status"] == "live"]
    planned = [r for r in results if r["status"] == "planned"]
    print(f"\n== Run summary — {len(live)} live · {len(planned)} planned · {len(failed)} failed ==")
    for r in results:
        marker = "OK  " if r["status"] == "live" else "FAIL" if r["status"] == "failed" else "SKIP"
        print(f"  {marker}  {r['id']:32s}  {r['rows']:>6} rows  {r['bytes']:>8} bytes")
    # Only workflow failure if at least one source was implemented AND failed.
    return 2 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
